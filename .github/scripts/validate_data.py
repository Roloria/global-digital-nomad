#!/usr/bin/env python3
"""数据校验脚本 — 在 GitHub Actions 中运行。

校验内容：
1. CSV 文件存在且格式正确（UTF-8，24 列）
2. 表头与 schema 一致
3. 评分数值在 0–10 范围
4. 必填字段无空值
5. 综合分计算结果与 CSV 中存储的值一致
6. CSV 与 Excel 数据一致（如果有 Excel 文件）
"""
import csv
import sys
import os
from pathlib import Path

EXPECTED_HEADERS = [
    "排名", "区域", "城市", "国家", "国家(英)", "国旗",
    "游民数", "月成本 (USD)", "货币类型",
    "网络", "社群", "生活", "安全",
    "英语", "步行", "空气", "女性友好", "LGBTQ+",
    "夜生活", "安静指数", "种族包容",
    "年均气温 (°C)", "最佳季节", "签证",
    "综合分",
    "最后更新",
]

SCORE_COLS = {
    "网络", "社群", "生活", "安全",
    "英语", "步行", "空气", "女性友好", "LGBTQ+",
    "夜生活", "安静指数", "种族包容",
    "综合分",
}

WEIGHTS = {
    "网络": 1.0, "社群": 1.2, "生活": 1.0, "安全": 1.2,
    "英语": 0.9, "步行": 0.7, "空气": 0.7, "女性友好": 0.8,
    "LGBTQ+": 0.6, "夜生活": 0.5, "安静指数": 0.6, "种族包容": 0.6,
}
WSUM = sum(WEIGHTS.values())  # 9.8

VALID_REGIONS = {"亚洲", "欧洲", "拉美", "非洲", "中东"}


def compute_overall(row):
    s = sum(float(row[h]) * w for h, w in WEIGHTS.items())
    return round(s / WSUM, 1)


def load_csv(path):
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def validate_csv(path):
    errors = []
    warnings = []

    if not Path(path).exists():
        errors.append(f"❌ 缺少 CSV 文件: {path}")
        return errors, warnings, []

    rows = load_csv(path)

    if not rows:
        errors.append("❌ CSV 文件为空")
        return errors, warnings, rows

    # Check headers
    actual_headers = list(rows[0].keys())
    if actual_headers != EXPECTED_HEADERS:
        errors.append(
            f"❌ CSV 表头不匹配\n"
            f"   期望: {EXPECTED_HEADERS}\n"
            f"   实际: {actual_headers}"
        )
        return errors, warnings, rows

    print(f"✅ CSV 表头正确（{len(actual_headers)} 列）")

    # Check each row
    cities_seen = set()
    for i, row in enumerate(rows, start=2):
        city = row["城市"]
        country = row["国家"]

        # Check unique city+country
        key = f"{city}|{country}"
        if key in cities_seen:
            errors.append(f"❌ 第 {i} 行：城市+国家 重复：{city}, {country}")
        cities_seen.add(key)

        # Check region
        region = row["区域"]
        if region not in VALID_REGIONS:
            errors.append(f"❌ 第 {i} 行：区域 '{region}' 不合法（应为 {VALID_REGIONS}）")

        # Check required fields
        for col in EXPECTED_HEADERS:
            v = row[col].strip()
            if not v:
                errors.append(f"❌ 第 {i} 行 ({city}): 字段 '{col}' 为空")

        # Check score ranges
        for col in SCORE_COLS:
            try:
                val = float(row[col])
                if val < 0 or val > 10:
                    errors.append(
                        f"❌ 第 {i} 行 ({city}): '{col}' = {val} 超出 0–10 范围"
                    )
            except ValueError:
                errors.append(
                    f"❌ 第 {i} 行 ({city}): '{col}' 不是数字: '{row[col]}'"
                )

        # Check nomads/cost/temp are integers
        for col in ["游民数", "月成本 (USD)", "年均气温 (°C)"]:
            try:
                int(row[col])
            except ValueError:
                errors.append(
                    f"❌ 第 {i} 行 ({city}): '{col}' 不是整数: '{row[col]}'"
                )

        # Sanity bounds（量级防呆，不限制具体数值）
        for col, lo, hi in [("月成本 (USD)", 100, 10000), ("游民数", 0, 10_000_000), ("年均气温 (°C)", -20, 40)]:
            try:
                v = int(row[col])
                if v < lo or v > hi:
                    errors.append(
                        f"❌ 第 {i} 行 ({city}): '{col}' = {v} 超出合理区间 [{lo}, {hi}]"
                    )
            except ValueError:
                pass  # 非整数已由上方校验报错

        # Check last_updated is YYYY-MM-DD format
        last_updated = row.get("最后更新", "").strip()
        import re as _re
        if not last_updated:
            errors.append(f"❌ 第 {i} 行 ({city}): '最后更新' 为空")
        elif not _re.match(r"^\d{4}-\d{2}-\d{2}$", last_updated):
            errors.append(
                f"❌ 第 {i} 行 ({city}): '最后更新' 格式错误: '{last_updated}' (应为 YYYY-MM-DD)"
            )

        # Check overall score
        expected_overall = compute_overall(row)
        actual_overall = float(row["综合分"])
        if abs(expected_overall - actual_overall) > 0.05:
            errors.append(
                f"❌ 第 {i} 行 ({city}): 综合分不一致\n"
                f"   期望: {expected_overall} | 实际: {actual_overall}"
            )

    print(f"✅ 共 {len(rows)} 行，已检查评分范围与综合分公式")

    # 排名语义：排名 = 按游民数降序的 1..N 排列（与网站的自动重算逻辑一致）
    ranks = [r["排名"] for r in rows]
    try:
        rank_ints = [int(x) for x in ranks]
        if sorted(rank_ints) != list(range(1, len(rows) + 1)):
            errors.append("❌ 排名列不是 1..N 的完整排列（存在缺失/重复/越界）")
        else:
            by_nomads = sorted(rows, key=lambda r: -int(r["游民数"]))
            for expect_i, r in enumerate(by_nomads, start=1):
                if int(r["排名"]) != expect_i:
                    errors.append(
                        f"❌ 城市 {r['城市']}: 排名 {r['排名']} 与游民数降序位次 {expect_i} 不符"
                    )
            if not errors or all("排名" not in e and "位次" not in e for e in errors):
                print("✅ 排名与游民数降序一致")
    except ValueError:
        errors.append("❌ 排名列存在非整数值")

    # Print top 5
    print("\n📊 综合分 TOP 5:")
    sorted_rows = sorted(rows, key=lambda r: -float(r["综合分"]))
    for i, r in enumerate(sorted_rows[:5], 1):
        print(f"   {i}. {r['城市']:<12} ({r['国家']:<10}) - {r['综合分']}")

    # Regional distribution
    print("\n🌍 区域分布:")
    region_count = {}
    for r in rows:
        region_count[r["区域"]] = region_count.get(r["区域"], 0) + 1
    for region, count in sorted(region_count.items()):
        print(f"   {region}: {count} 城")

    return errors, warnings, rows


def validate_xlsx_matches_csv(csv_rows):
    """校验 Excel 与 CSV 数据一致（如果有 Excel 文件）。"""
    errors = []
    xlsx_path = os.environ.get("XLSX_PATH", "data/digital-nomad-cities.xlsx")
    if not Path(xlsx_path).exists():
        print(f"⚠️  Excel 文件不存在，跳过对比: {xlsx_path}")
        return errors

    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl 未安装，跳过 Excel 对比")
        return errors

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["综合数据"]

    # Find data start (skip title rows)
    header_row = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "排名":
            header_row = r_idx
            break

    if header_row is None:
        errors.append("❌ Excel 中找不到 '排名' 表头行")
        return errors

    xlsx_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=len(EXPECTED_HEADERS), values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(EXPECTED_HEADERS, row))
        xlsx_rows.append(d)

    if len(xlsx_rows) != len(csv_rows):
        errors.append(
            f"❌ 行数不一致: CSV={len(csv_rows)} vs Excel={len(xlsx_rows)}"
        )
        return errors

    # Compare each row
    for i, (csv_row, xlsx_row) in enumerate(zip(csv_rows, xlsx_rows)):
        for col in EXPECTED_HEADERS:
            csv_val = str(csv_row[col]).strip()
            xlsx_val = str(xlsx_row[col]).strip() if xlsx_row[col] is not None else ""
            if csv_val != xlsx_val:
                errors.append(
                    f"❌ 第 {i+1} 行 ({csv_row['城市']}): 字段 '{col}' 不一致\n"
                    f"   CSV: '{csv_val}' | Excel: '{xlsx_val}'"
                )

    if not errors:
        print(f"✅ Excel 与 CSV 数据一致（{len(xlsx_rows)} 行 × {len(EXPECTED_HEADERS)} 列）")

    return errors


def main():
    csv_path = os.environ.get("CSV_PATH", "data/digital-nomad-cities.csv")
    print(f"🔍 校验数据文件: {csv_path}\n")

    errors, warnings, csv_rows = validate_csv(csv_path)

    if csv_rows:
        errors.extend(validate_xlsx_matches_csv(csv_rows))

    print("\n" + "=" * 50)
    if errors:
        print(f"\n❌ 校验失败：发现 {len(errors)} 个错误\n")
        for e in errors:
            print(e)
            print()
        sys.exit(1)
    else:
        print("\n✅ 所有校验通过！")


if __name__ == "__main__":
    main()
