#!/usr/bin/env python3
"""数字游民社区数据集校验脚本 — 在 GitHub Actions 中运行。

校验内容：
1. CSV 文件存在且格式正确（UTF-8，18 列）
2. 表头与 schema 一致
3. 必填字段无空值
4. 评分在 0–10 范围
5. URL/邮箱格式基本校验
6. Excel 与 CSV 数据一致（如果有 Excel 文件）
"""
import csv
import sys
import os
import re
from pathlib import Path

EXPECTED_HEADERS = [
    "排名", "区域", "城市", "国家", "国旗",
    "社区名称", "社区名称(英)", "类型", "简介",
    "月费(USD)", "容量", "政策摘要",
    "网址", "联系邮箱", "社群链接",
    "综合分", "来源", "最后更新",
]

VALID_TYPES = {"联合办公", "联合生活", "聚会", "度假村", "在线社群"}
VALID_REGIONS = {"亚洲", "欧洲", "拉美", "非洲", "中东", "全球"}

URL_RE = re.compile(r"^https?://[^\s]+|^-$|^$")
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$|^-$")


def load_csv(path):
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def validate_csv(path):
    errors, warnings = [], []

    if not Path(path).exists():
        errors.append(f"❌ 缺少 CSV 文件: {path}")
        return errors, warnings, []

    rows = load_csv(path)

    if not rows:
        errors.append("❌ CSV 文件为空")
        return errors, warnings, rows

    actual_headers = list(rows[0].keys())
    if actual_headers != EXPECTED_HEADERS:
        errors.append(
            f"❌ CSV 表头不匹配\n"
            f"   期望: {EXPECTED_HEADERS}\n"
            f"   实际: {actual_headers}"
        )
        return errors, warnings, rows

    print(f"✅ CSV 表头正确（{len(actual_headers)} 列）")

    seen = set()
    for i, row in enumerate(rows, start=2):
        name = row["社区名称"]
        city = row["城市"]
        country = row["国家"]

        key = f"{name}|{city}|{country}"
        if key in seen:
            errors.append(f"❌ 第 {i} 行：社区+城市+国家 重复：{name} @ {city}, {country}")
        seen.add(key)

        if row["区域"] not in VALID_REGIONS:
            errors.append(f"❌ 第 {i} 行 ({name}): 区域 '{row['区域']}' 不合法（应为 {VALID_REGIONS}）")

        if row["类型"] not in VALID_TYPES:
            errors.append(f"❌ 第 {i} 行 ({name}): 类型 '{row['类型']}' 不合法（应为 {VALID_TYPES}）")

        for col in EXPECTED_HEADERS:
            v = row[col].strip()
            if not v:
                errors.append(f"❌ 第 {i} 行 ({name}): 字段 '{col}' 为空")

        try:
            score = float(row["综合分"])
            if score < 0 or score > 10:
                errors.append(f"❌ 第 {i} 行 ({name}): 综合分 {score} 超出 0–10 范围")
        except ValueError:
            errors.append(f"❌ 第 {i} 行 ({name}): 综合分不是数字: '{row['综合分']}'")

        try:
            int(row["容量"])
        except ValueError:
            errors.append(f"❌ 第 {i} 行 ({name}): 容量不是整数: '{row['容量']}'")

        lu = row.get("最后更新", "").strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", lu):
            errors.append(f"❌ 第 {i} 行 ({name}): '最后更新' 格式错误: '{lu}'")

        if not URL_RE.match(row["网址"]):
            errors.append(f"❌ 第 {i} 行 ({name}): 网址格式错误: '{row['网址']}'")

        if not EMAIL_RE.match(row["联系邮箱"]):
            errors.append(f"❌ 第 {i} 行 ({name}): 联系邮箱格式错误: '{row['联系邮箱']}'")

        if not URL_RE.match(row["社群链接"]):
            errors.append(f"❌ 第 {i} 行 ({name}): 社群链接格式错误: '{row['社群链接']}'")

    print(f"✅ 共 {len(rows)} 行，已检查类型 / 区域 / 评分 / URL / 邮箱")

    print("\n🌟 综合分 TOP 10:")
    for i, r in enumerate(sorted(rows, key=lambda x: -float(x["综合分"]))[:10], 1):
        print(f"   {i:>2}. {r['社区名称']:<22} ({r['城市']}, {r['国家']:<10}) - {r['综合分']} · {r['类型']}")

    from collections import Counter
    print("\n📂 类型分布:")
    for k, v in Counter(r["类型"] for r in rows).most_common():
        print(f"   {k}: {v}")
    print("\n🌍 区域分布:")
    for k, v in Counter(r["区域"] for r in rows).most_common():
        print(f"   {k}: {v}")

    return errors, warnings, rows


def validate_xlsx_matches_csv(csv_rows):
    errors = []
    xlsx_path = os.environ.get("XLSX_PATH", "data/digital-nomad-communities.xlsx")
    if not Path(xlsx_path).exists():
        print(f"⚠️  Excel 文件不存在，跳过对比: {xlsx_path}")
        return errors

    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl 未安装，跳过 Excel 对比")
        return errors

    wb = openpyxl.load_workbook(xlsx_path)

    main_sheet = None
    for name in wb.sheetnames:
        if "综合" in name or "main" in name.lower() or len(wb.sheetnames) == 1:
            main_sheet = wb[name]
            break
    if not main_sheet:
        main_sheet = wb[wb.sheetnames[0]]

    header_row = None
    for r_idx, row in enumerate(main_sheet.iter_rows(values_only=True), start=1):
        if row and row[0] == "排名":
            header_row = r_idx
            break

    if header_row is None:
        errors.append("❌ Excel 中找不到 '排名' 表头行")
        return errors

    xlsx_rows = []
    for row in main_sheet.iter_rows(min_row=header_row + 1, max_col=len(EXPECTED_HEADERS), values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(EXPECTED_HEADERS, row))
        xlsx_rows.append(d)

    if len(xlsx_rows) != len(csv_rows):
        errors.append(f"❌ 行数不一致: CSV={len(csv_rows)} vs Excel={len(xlsx_rows)}")
        return errors

    def norm(v):
        """Normalize: 8 == 8.0; strip whitespace."""
        if v is None:
            return ""
        s = str(v).strip()
        # Try float compare
        try:
            f = float(s)
            # Check if integer float (8.0 -> 8 but stored as 8.0 in CSV)
            if f.is_integer() and "." in str(csv_rows[0].get(EXPECTED_HEADERS[0], "")):
                # keep as is
                pass
            return s
        except (ValueError, TypeError):
            return s

    for i, (csv_row, xlsx_row) in enumerate(zip(csv_rows, xlsx_rows)):
        for col in EXPECTED_HEADERS:
            csv_val = str(csv_row[col]).strip()
            xlsx_val = str(xlsx_row[col]).strip() if xlsx_row[col] is not None else ""
            # Try numeric comparison
            try:
                if float(csv_val) == float(xlsx_val):
                    continue
            except (ValueError, TypeError):
                pass
            if csv_val != xlsx_val:
                errors.append(
                    f"❌ 第 {i+1} 行 ({csv_row['社区名称']}): 字段 '{col}' 不一致\n"
                    f"   CSV: '{csv_val}' | Excel: '{xlsx_val}'"
                )

    if not errors:
        print(f"✅ Excel 与 CSV 数据一致（{len(xlsx_rows)} 行 × {len(EXPECTED_HEADERS)} 列）")
    return errors


def main():
    csv_path = os.environ.get("CSV_PATH", "data/digital-nomad-communities.csv")
    print(f"🔍 校验数字游民社区数据文件: {csv_path}\n")

    errors, warnings, csv_rows = validate_csv(csv_path)

    if csv_rows:
        errors.extend(validate_xlsx_matches_csv(csv_rows))

    print("\n" + "=" * 50)
    if errors:
        print(f"\n❌ 校验失败：发现 {len(errors)} 个错误\n")
        for e in errors:
            print(e)
            print()
        sys.exit(1)
    else:
        print("\n✅ 所有校验通过！")


if __name__ == "__main__":
    main()
